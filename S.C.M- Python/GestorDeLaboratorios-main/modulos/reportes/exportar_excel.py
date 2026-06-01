from flask import Blueprint, request, send_file, make_response
import traceback
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from io import BytesIO
from config.db import get_connection

bp = Blueprint('exportar_excel', __name__)

@bp.route('/exportar_excel_tarjetas')
def exportar_excel_tarjetas():
    try:
        fecha_desde = request.args.get('fecha_desde', '')
        fecha_hasta = request.args.get('fecha_hasta', '')
        persona = request.args.get('persona', '')
        usuario = request.args.get('usuario', '')
        tarjeta = request.args.get('tarjeta', '')

        query = "SELECT * FROM historial_tarjetas WHERE 1=1"
        params = []

        if fecha_desde:
            query += " AND fecha_hora >= %s"
            params.append(f"{fecha_desde} 00:00:00")
        if fecha_hasta:
            query += " AND fecha_hora <= %s"
            params.append(f"{fecha_hasta} 23:59:59")
        if persona:
            query += " AND nombre_completo LIKE %s"
            params.append(f"%{persona}%")
        if usuario:
            query += " AND ejecutado_por LIKE %s"
            params.append(f"%{usuario}%")
        if tarjeta:
            query += " AND uid LIKE %s"
            params.append(f"%{tarjeta}%")

        query += " ORDER BY fecha_hora DESC"

        connection = get_connection()
        try:
            cursor = connection.cursor()
            try:
                exec_query = query
                if connection.__class__.__module__.startswith('sqlite3'):
                    exec_query = exec_query.replace('%s', '?')
                cursor.execute(exec_query, params)
                resultados = cursor.fetchall()
                # Normalizar filas a diccionarios independientemente del driver
                if resultados:
                    if hasattr(resultados[0], 'keys'):
                        resultados = [dict(r) for r in resultados]
                    else:
                        cols = [c[0] for c in cursor.description] if cursor.description else []
                        resultados = [dict(zip(cols, r)) for r in resultados]
            finally:
                try:
                    cursor.close()
                except Exception:
                    pass
        finally:
            connection.close()

        wb = Workbook()
        ws = wb.active
        ws.title = "Historial Registro Personas"

        headers = ['ID Reg.', 'Tarjeta UID / Credencial', 'Persona Registrada', 'Acción Realizada', 'Usuario Autor / Responsable', 'Fecha de Registro']
        ws.append(headers)

        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        for row in resultados:
            # Mapear acción a formato legible
            accion_mapeada = row.get('accion', '')
            if accion_mapeada == 'alta':
                accion_mapeada = 'Alta'
            elif accion_mapeada == 'baja':
                accion_mapeada = 'Baja'
            elif accion_mapeada == 'edicion':
                accion_mapeada = 'Edición'

            ws.append([
                row.get('id'),
                row.get('uid') or row.get('uid_tarjeta') or '',
                row.get('nombre_completo') or '',
                accion_mapeada,
                row.get('ejecutado_por') or '',
                row.get('fecha_hora') or ''
            ])

        for column in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column)
            column_letter = column[0].column_letter
            ws.column_dimensions[column_letter].width = max_length + 2

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            download_name="historial_altas_bajas_tarjetas.xlsx",
            as_attachment=True,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception:
        tb = traceback.format_exc()
        return make_response(tb, 500)

@bp.route('/exportar_excel_accesos')
def exportar_excel_accesos():
    fecha_desde = request.args.get('fecha_desde', '')
    fecha_hasta = request.args.get('fecha_hasta', '')
    tipo_movimiento = request.args.get('tipo_movimiento', '')
    resultado = request.args.get('resultado', '')
    credencial = request.args.get('credencial', '')
    nombre_completo = request.args.get('nombre_completo', '')
    uid_tarjeta = request.args.get('uid_tarjeta', '')

    query = '''
    SELECT
        ra.id,
        ra.fecha_hora,
        COALESCE(p.nombre_completo, 'Desconocido') AS persona,
        CASE
            WHEN LOWER(TRIM(tm.movimiento)) = 'entrada' THEN 'Entrada'
            WHEN LOWER(TRIM(tm.movimiento)) = 'salida' THEN 'Salida'
            WHEN LOWER(ra.descripcion) LIKE '%retirar%' THEN 'Entrada'
            WHEN LOWER(ra.descripcion) LIKE '%devolver%' THEN 'Salida'
            ELSE COALESCE(tm.movimiento, ra.descripcion, 'Desconocido')
        END AS movimiento,
        ra.resultado,
        CASE
            WHEN LOWER(TRIM(COALESCE(ra.credencial, 'Tarjeta'))) IN ('uid', 'tarjeta') THEN 'TARJETA'
            WHEN LOWER(TRIM(ra.credencial)) = 'pin' THEN 'PIN'
            ELSE UPPER(ra.credencial)
        END AS credencial,
        ra.tarjeta_uid,
        ra.accion,
        ra.descripcion
    FROM registro_acceso ra
    LEFT JOIN enrolar e ON ra.enrolar_id = e.id
    LEFT JOIN personas p ON e.persona_id = p.id
    LEFT JOIN tipo_movimiento tm ON ra.tipo_movimiento_id = tm.id
    WHERE 1=1
    '''
    params = []

    if fecha_desde:
        query += " AND ra.fecha_hora >= %s"
        params.append(f"{fecha_desde} 00:00:00")
    if fecha_hasta:
        query += " AND ra.fecha_hora <= %s"
        params.append(f"{fecha_hasta} 23:59:59")
    if tipo_movimiento:
        query += " AND LOWER(TRIM(tm.movimiento)) = %s"
        params.append(tipo_movimiento.strip().lower())
    if resultado:
        query += " AND ra.resultado = %s"
        params.append(resultado)
    if credencial:
        credencial_val = credencial.strip().lower()
        if credencial_val == 'tarjeta':
            query += " AND LOWER(TRIM(ra.credencial)) IN ('tarjeta', 'uid')"
        elif credencial_val == 'pin':
            query += " AND LOWER(TRIM(ra.credencial)) = 'pin'"
        else:
            query += " AND LOWER(TRIM(ra.credencial)) = %s"
            params.append(credencial_val)
    if nombre_completo:
        query += " AND p.nombre_completo LIKE %s"
        params.append(f"%{nombre_completo}%")
    if uid_tarjeta:
        query += " AND ra.tarjeta_uid LIKE %s"
        params.append(f"%{uid_tarjeta}%")

    query += " ORDER BY ra.fecha_hora DESC"

    connection = get_connection()
    try:
        cursor = connection.cursor()
        try:
            exec_query = query
            if connection.__class__.__module__.startswith('sqlite3'):
                exec_query = exec_query.replace('%s', '?')
            cursor.execute(exec_query, params)
            resultados = cursor.fetchall()
            if resultados:
                if hasattr(resultados[0], 'keys'):
                    resultados = [dict(r) for r in resultados]
                else:
                    cols = [c[0] for c in cursor.description] if cursor.description else []
                    resultados = [dict(zip(cols, r)) for r in resultados]
        finally:
            try:
                cursor.close()
            except Exception:
                pass
    finally:
        connection.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Accesos"

    headers = ['ID', 'Fecha y Hora', 'Persona', 'Movimiento', 'Acción ESP', 'Resultado', 'Credencial', 'UID Tarjeta', 'Descripción']
    ws.append(headers)

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for row in resultados:
        ws.append([
            row.get('id'),
            row.get('fecha_hora'),
            row.get('persona'),
            row.get('movimiento'),
            row.get('accion') or '-',
            row.get('resultado'),
            row.get('credencial'),
            row.get('tarjeta_uid'),
            row.get('descripcion')
        ])

    for column in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column)
        column_letter = column[0].column_letter
        ws.column_dimensions[column_letter].width = max_length + 2

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        download_name="reporte_accesos.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@bp.route('/exportar_excel_enrolamiento')
def exportar_excel_enrolamiento():
    fecha_desde = request.args.get('fecha_desde', '')
    fecha_hasta = request.args.get('fecha_hasta', '')
    accion = request.args.get('accion', '')
    responsable = request.args.get('responsable', '')
    persona = request.args.get('persona', '')
    tarjeta = request.args.get('tarjeta', '')

    query = '''
    SELECT
        id,
        enrolar_id,
        persona_id,
        nombre_persona,
        perfil,
        tarjeta_uid,
        tarjeta_pin,
        estado,
        responsable,
        descripcion,
        fecha_hora
    FROM historial_enrolamiento
    WHERE 1=1
    '''
    params = []

    if fecha_desde:
        query += " AND fecha_hora >= %s"
        params.append(f"{fecha_desde} 00:00:00")
    if fecha_hasta:
        query += " AND fecha_hora <= %s"
        params.append(f"{fecha_hasta} 23:59:59")
    if accion:
        query += " AND LOWER(TRIM(estado)) = %s"
        params.append(accion.strip().lower())
    if responsable:
        query += " AND LOWER(responsable) LIKE %s"
        params.append(f"%{responsable.strip().lower()}%")
    if persona:
        query += " AND LOWER(nombre_persona) LIKE %s"
        params.append(f"%{persona.strip().lower()}%")
    if tarjeta:
        query += " AND (LOWER(tarjeta_uid) LIKE %s OR LOWER(tarjeta_pin) LIKE %s)"
        params.append(f"%{tarjeta.strip().lower()}%")
        params.append(f"%{tarjeta.strip().lower()}%")

    query += " ORDER BY fecha_hora DESC"

    connection = get_connection()
    try:
        cursor = connection.cursor()
        try:
            exec_query = query
            if connection.__class__.__module__.startswith('sqlite3'):
                exec_query = exec_query.replace('%s', '?')
            cursor.execute(exec_query, params)
            resultados = cursor.fetchall()
            if resultados and hasattr(resultados[0], 'keys'):
                resultados = [dict(r) for r in resultados]
        finally:
            try:
                cursor.close()
            except Exception:
                pass
    finally:
        connection.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Historial Enrolamiento"

    headers = ['ID', 'Enrolar ID', 'Persona', 'Perfil', 'Tarjeta UID', 'PIN', 'Estado', 'Responsable', 'Descripción', 'Fecha y Hora']
    ws.append(headers)

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for row in resultados:
        ws.append([
            row.get('id'),
            row.get('enrolar_id'),
            row.get('nombre_persona'),
            row.get('perfil'),
            row.get('tarjeta_uid'),
            row.get('tarjeta_pin'),
            row.get('estado'),
            row.get('responsable'),
            row.get('descripcion'),
            row.get('fecha_hora')
        ])

    for column in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column)
        column_letter = column[0].column_letter
        ws.column_dimensions[column_letter].width = max_length + 2

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        download_name="historial_enrolamiento.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
